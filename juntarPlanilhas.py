import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, NamedStyle

def juntarPlanilhas():
    tabelaOrganizada = pd.read_excel("W:\\Laboratorio\\TUBOS\\Diversos\\Programas\\Arquivos_temp\\Resultado.xlsx")
    tabelaOriginal = pd.read_excel("W:\\Laboratorio\\TUBOS\\Diversos\\Planilha de OP's.xlsx", header=2)
    tabelaOriginal2 = pd.read_excel("W:\\Laboratorio\\TUBOS\\Diversos\\Planilha de OP's.xlsx", sheet_name=1)
    tabelaOriginal = tabelaOriginal.set_axis(['OP', 'Material', 'Código Produto', 'Tubo Fundido', '% Cu Mín', '% Cu Máx', '% P Mín', '% P Máx'], axis=1)
    tabelaOriginal.dropna(axis=1, how='all', inplace=True)
    tabelaOriginal2.dropna(axis=1, how='all', inplace=True)
    tabelaOrganizada.dropna(axis=1, how='all', inplace=True)

    tabelaFinal = pd.concat([tabelaOriginal, tabelaOrganizada])
    tabelaFinal = tabelaFinal.drop_duplicates()

    numeros_repetidos = set(tabelaFinal['OP']).intersection(tabelaOriginal2['OP'])

    tabelaFinal = tabelaFinal[~tabelaFinal['OP'].isin(numeros_repetidos)]

    tabelaFinal['Ano'] = tabelaFinal['OP'].str.extract(r'^.{2}(.{2})')
    tabelaFinal['Semana'] = tabelaFinal['OP'].str.extract(r'(^.{2})')
    tabelaFinal['Lote'] = tabelaFinal['OP'].str.extract(r'^.{2}.{2}-(.{6})')

    tabelaFinal = tabelaFinal.sort_values(['Ano', 'Semana', 'Lote'])
    tabelaFinal = tabelaFinal.drop(columns=['Ano', 'Semana', 'Lote'])

    tabelaFinal.to_excel('W:\\Laboratorio\\TUBOS\\Diversos\\Programas\\Arquivos_temp\\temp1.xlsx', index=False)

    tabelaFinal = openpyxl.load_workbook('W:\\Laboratorio\\TUBOS\\Diversos\\Programas\\Arquivos_temp\\temp1.xlsx')
    sheetFinal = tabelaFinal.active

    tabelaOriginal = openpyxl.load_workbook("W:\\Laboratorio\\TUBOS\\Diversos\\Planilha de OP's.xlsx")
    sheetOriginal = tabelaOriginal.active

    max_row_final = sheetFinal.max_row
    max_column_final = sheetFinal.max_column

    linha_destino = 4

    for linha_origem in sheetFinal.iter_rows(min_row=2, max_row=max_row_final, values_only=True):
        for coluna_origem, valor in enumerate(linha_origem, start=1):
            sheetOriginal.cell(row=linha_destino, column=coluna_origem, value=valor)
        linha_destino += 1

    numero_style = NamedStyle(name='numero_style', number_format='0.000')

    for coluna_letra in ['G', 'H']:
        for row in range(4, max_row_final + 4):
            sheetOriginal[coluna_letra + str(row)].number_format = numero_style.number_format

    for row in range(4, max_row_final + 4):
        for column in range(1, max_column_final + 1):
            cell = sheetOriginal.cell(row=row, column=column)

            alignment = Alignment(horizontal='center', vertical='center')

            border_style = Side(style='thin')

            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

            cell.alignment = alignment
            cell.border = border

    tabelaOriginal.save("W:\\Laboratorio\\TUBOS\\Diversos\\Planilha de OP's.xlsx")

    tabelaFinal.close()
    tabelaOriginal.close()
    os.remove("W:\\Laboratorio\\TUBOS\\Diversos\\Programas\\Arquivos_temp\\temp1.xlsx")
    os.remove("W:\\Laboratorio\\TUBOS\\Diversos\\Programas\\Arquivos_temp\\Resultado.xlsx")